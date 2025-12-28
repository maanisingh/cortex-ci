"""
Compliance Reporting Engine
Uses open source tools only:
- WeasyPrint for PDF generation (HTML -> PDF)
- openpyxl for Excel generation
- Jinja2 for templating

No external services or sign-ups required.
"""

import io
from datetime import datetime
from pathlib import Path
from typing import Any

# PDF Generation (WeasyPrint)
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

# Excel Generation (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# HTML templating
from jinja2 import Environment, BaseLoader


class ComplianceReportGenerator:
    """Generate compliance reports in PDF and Excel formats."""

    def __init__(self):
        self.jinja_env = Environment(loader=BaseLoader())

    # ============ PDF GENERATION ============

    def generate_compliance_report_pdf(
        self,
        company_data: dict[str, Any],
        compliance_data: dict[str, Any],
        framework: str = "all",
    ) -> bytes:
        """Generate a comprehensive compliance report as PDF."""
        html_content = self._render_compliance_report_html(
            company_data, compliance_data, framework
        )

        if not WEASYPRINT_AVAILABLE:
            # Fallback: return HTML if WeasyPrint not installed
            return html_content.encode("utf-8")

        # Generate PDF
        html = HTML(string=html_content)
        css = CSS(string=self._get_report_css())
        pdf_bytes = html.write_pdf(stylesheets=[css])
        return pdf_bytes

    def _render_compliance_report_html(
        self,
        company_data: dict[str, Any],
        compliance_data: dict[str, Any],
        framework: str,
    ) -> str:
        """Render compliance report as HTML."""
        template_str = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Отчёт о соответствии — {{ company.name }}</title>
</head>
<body>
    <div class="header">
        <h1>ОТЧЁТ О СООТВЕТСТВИИ</h1>
        <h2>{{ company.name }}</h2>
        <p>ИНН: {{ company.inn }} | Дата: {{ report_date }}</p>
    </div>

    <div class="summary">
        <h2>1. Сводка</h2>
        <table class="summary-table">
            <tr>
                <td class="label">Общий уровень соответствия:</td>
                <td class="value score-{{ score_class }}">{{ compliance.overall_score }}%</td>
            </tr>
            <tr>
                <td class="label">Проверено требований:</td>
                <td class="value">{{ compliance.total_checks }}</td>
            </tr>
            <tr>
                <td class="label">Выполнено:</td>
                <td class="value text-green">{{ compliance.passed }}</td>
            </tr>
            <tr>
                <td class="label">Частично выполнено:</td>
                <td class="value text-yellow">{{ compliance.warnings }}</td>
            </tr>
            <tr>
                <td class="label">Не выполнено:</td>
                <td class="value text-red">{{ compliance.failed }}</td>
            </tr>
        </table>
    </div>

    <div class="frameworks">
        <h2>2. Соответствие по нормативным актам</h2>
        {% for fw in frameworks %}
        <div class="framework-section">
            <h3>{{ fw.name }}</h3>
            <div class="progress-bar">
                <div class="progress" style="width: {{ fw.score }}%"></div>
            </div>
            <p class="score">{{ fw.score }}% ({{ fw.completed }}/{{ fw.total }} требований)</p>
        </div>
        {% endfor %}
    </div>

    <div class="details">
        <h2>3. Детализация по категориям</h2>
        {% for category in categories %}
        <div class="category-section">
            <h3>{{ category.name }}</h3>
            <table class="checks-table">
                <thead>
                    <tr>
                        <th>Требование</th>
                        <th>Статус</th>
                        <th>Комментарий</th>
                    </tr>
                </thead>
                <tbody>
                    {% for item in category.items %}
                    <tr class="status-{{ item.status }}">
                        <td>{{ item.check }}</td>
                        <td class="status">
                            {% if item.status == 'pass' %}✓{% elif item.status == 'warning' %}⚠{% else %}✗{% endif %}
                        </td>
                        <td>{{ item.comment }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% endfor %}
    </div>

    <div class="recommendations">
        <h2>4. Рекомендации</h2>
        <ol>
            {% for rec in recommendations %}
            <li class="priority-{{ rec.priority }}">
                <strong>[{{ rec.priority | upper }}]</strong> {{ rec.action }}
            </li>
            {% endfor %}
        </ol>
    </div>

    <div class="documents">
        <h2>5. Состояние документации</h2>
        <table class="docs-table">
            <thead>
                <tr>
                    <th>Документ</th>
                    <th>Статус</th>
                    <th>Дата</th>
                </tr>
            </thead>
            <tbody>
                {% for doc in documents %}
                <tr>
                    <td>{{ doc.title }}</td>
                    <td class="status-{{ doc.status }}">{{ doc.status_label }}</td>
                    <td>{{ doc.date }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>

    <div class="footer">
        <p>Отчёт сформирован автоматически системой Cortex GRC</p>
        <p>{{ report_date }} | {{ company.name }}</p>
    </div>
</body>
</html>
        """

        template = self.jinja_env.from_string(template_str)

        # Prepare data
        score = compliance_data.get("overall_score", 0)
        score_class = "green" if score >= 80 else "yellow" if score >= 60 else "red"

        return template.render(
            company=company_data,
            compliance=compliance_data,
            report_date=datetime.now().strftime("%d.%m.%Y"),
            score_class=score_class,
            frameworks=compliance_data.get("frameworks", []),
            categories=compliance_data.get("categories", []),
            recommendations=compliance_data.get("recommendations", []),
            documents=compliance_data.get("documents", []),
        )

    def _get_report_css(self) -> str:
        """Get CSS styles for PDF report."""
        return """
        @page {
            size: A4;
            margin: 2cm;
        }
        body {
            font-family: 'DejaVu Sans', 'Arial', sans-serif;
            font-size: 10pt;
            line-height: 1.4;
            color: #333;
        }
        .header {
            text-align: center;
            border-bottom: 2px solid #2563eb;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        .header h1 {
            color: #2563eb;
            font-size: 24pt;
            margin: 0;
        }
        .header h2 {
            font-size: 16pt;
            margin: 10px 0;
        }
        h2 {
            color: #1e40af;
            border-bottom: 1px solid #ddd;
            padding-bottom: 5px;
            margin-top: 30px;
        }
        h3 {
            color: #374151;
            margin-top: 20px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 10px 0;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        th {
            background-color: #f3f4f6;
            font-weight: bold;
        }
        .summary-table {
            width: 50%;
        }
        .summary-table td {
            border: none;
            padding: 5px 10px;
        }
        .summary-table .label {
            font-weight: bold;
        }
        .summary-table .value {
            text-align: right;
            font-size: 14pt;
        }
        .score-green { color: #059669; }
        .score-yellow { color: #d97706; }
        .score-red { color: #dc2626; }
        .text-green { color: #059669; }
        .text-yellow { color: #d97706; }
        .text-red { color: #dc2626; }
        .progress-bar {
            width: 100%;
            height: 20px;
            background-color: #e5e7eb;
            border-radius: 10px;
            overflow: hidden;
        }
        .progress {
            height: 100%;
            background-color: #2563eb;
        }
        .status-pass { background-color: #d1fae5; }
        .status-warning { background-color: #fef3c7; }
        .status-fail { background-color: #fee2e2; }
        .priority-high { color: #dc2626; }
        .priority-medium { color: #d97706; }
        .priority-low { color: #6b7280; }
        .footer {
            margin-top: 50px;
            text-align: center;
            font-size: 8pt;
            color: #9ca3af;
            border-top: 1px solid #ddd;
            padding-top: 10px;
        }
        """

    # ============ EXCEL GENERATION ============

    def generate_compliance_report_excel(
        self,
        company_data: dict[str, Any],
        compliance_data: dict[str, Any],
        framework: str = "all",
    ) -> bytes:
        """Generate compliance report as Excel file."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")

        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        self._create_summary_sheet(ws_summary, company_data, compliance_data)

        # Sheet 2: Detailed Checks
        ws_details = wb.create_sheet("Детализация")
        self._create_details_sheet(ws_details, compliance_data)

        # Sheet 3: Documents
        ws_docs = wb.create_sheet("Документы")
        self._create_documents_sheet(ws_docs, compliance_data)

        # Sheet 4: Recommendations
        ws_recs = wb.create_sheet("Рекомендации")
        self._create_recommendations_sheet(ws_recs, compliance_data)

        # Sheet 5: Measures Matrix
        ws_measures = wb.create_sheet("Матрица мер")
        self._create_measures_sheet(ws_measures, compliance_data)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(
        self,
        ws,
        company_data: dict[str, Any],
        compliance_data: dict[str, Any],
    ):
        """Create summary sheet in Excel."""
        # Styles
        title_font = Font(size=16, bold=True, color="1E40AF")
        header_font = Font(size=12, bold=True)
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Title
        ws["A1"] = "ОТЧЁТ О СООТВЕТСТВИИ"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        ws["A2"] = company_data.get("name", "")
        ws["A2"].font = Font(size=14, bold=True)
        ws.merge_cells("A2:D2")

        ws["A3"] = f"ИНН: {company_data.get('inn', '')} | Дата: {datetime.now().strftime('%d.%m.%Y')}"
        ws.merge_cells("A3:D3")

        # Summary section
        ws["A5"] = "СВОДКА"
        ws["A5"].font = header_font

        summary_data = [
            ("Общий уровень соответствия", f"{compliance_data.get('overall_score', 0)}%"),
            ("Проверено требований", compliance_data.get("total_checks", 0)),
            ("Выполнено", compliance_data.get("passed", 0)),
            ("Частично выполнено", compliance_data.get("warnings", 0)),
            ("Не выполнено", compliance_data.get("failed", 0)),
        ]

        for i, (label, value) in enumerate(summary_data, start=6):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Frameworks section
        row = 12
        ws[f"A{row}"] = "СООТВЕТСТВИЕ ПО НОРМАТИВНЫМ АКТАМ"
        ws[f"A{row}"].font = header_font
        row += 1

        headers = ["Нормативный акт", "Уровень соответствия", "Выполнено", "Всего"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        row += 1

        for fw in compliance_data.get("frameworks", []):
            ws.cell(row=row, column=1, value=fw.get("name", ""))
            ws.cell(row=row, column=2, value=f"{fw.get('score', 0)}%")
            ws.cell(row=row, column=3, value=fw.get("completed", 0))
            ws.cell(row=row, column=4, value=fw.get("total", 0))
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_details_sheet(self, ws, compliance_data: dict[str, Any]):
        """Create detailed checks sheet."""
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        headers = ["Категория", "Требование", "Статус", "Комментарий"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        row = 2
        for category in compliance_data.get("categories", []):
            for item in category.get("items", []):
                ws.cell(row=row, column=1, value=category.get("name", ""))
                ws.cell(row=row, column=2, value=item.get("check", ""))

                status = item.get("status", "")
                status_cell = ws.cell(row=row, column=3)
                if status == "pass":
                    status_cell.value = "Выполнено"
                    status_cell.fill = pass_fill
                elif status == "warning":
                    status_cell.value = "Частично"
                    status_cell.fill = warn_fill
                else:
                    status_cell.value = "Не выполнено"
                    status_cell.fill = fail_fill

                ws.cell(row=row, column=4, value=item.get("comment", ""))
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40

    def _create_documents_sheet(self, ws, compliance_data: dict[str, Any]):
        """Create documents status sheet."""
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        headers = ["Документ", "Статус", "Дата создания", "Дата утверждения", "Ответственный"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        row = 2
        for doc in compliance_data.get("documents", []):
            ws.cell(row=row, column=1, value=doc.get("title", ""))
            ws.cell(row=row, column=2, value=doc.get("status_label", ""))
            ws.cell(row=row, column=3, value=doc.get("created_at", ""))
            ws.cell(row=row, column=4, value=doc.get("approved_at", ""))
            ws.cell(row=row, column=5, value=doc.get("responsible", ""))
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25

    def _create_recommendations_sheet(self, ws, compliance_data: dict[str, Any]):
        """Create recommendations sheet."""
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        medium_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        headers = ["№", "Приоритет", "Рекомендация", "Срок"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        row = 2
        for i, rec in enumerate(compliance_data.get("recommendations", []), start=1):
            ws.cell(row=row, column=1, value=i)

            priority = rec.get("priority", "low")
            priority_cell = ws.cell(row=row, column=2)
            if priority == "high":
                priority_cell.value = "Высокий"
                priority_cell.fill = high_fill
            elif priority == "medium":
                priority_cell.value = "Средний"
                priority_cell.fill = medium_fill
            else:
                priority_cell.value = "Низкий"

            ws.cell(row=row, column=3, value=rec.get("action", ""))
            ws.cell(row=row, column=4, value=rec.get("deadline", ""))
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 15

    def _create_measures_sheet(self, ws, compliance_data: dict[str, Any]):
        """Create security measures matrix sheet."""
        header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        headers = ["Код", "Мера защиты", "УЗ-4", "УЗ-3", "УЗ-2", "УЗ-1", "Статус", "Средство реализации"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Example measures (would come from compliance_data in production)
        measures = [
            ("ИАФ.1", "Идентификация и аутентификация пользователей", "+", "+", "+", "+", "Реализовано", "Active Directory"),
            ("УПД.1", "Управление учётными записями", "+", "+", "+", "+", "Реализовано", "AD GPO"),
            ("АВЗ.1", "Антивирусная защита", "+", "+", "+", "+", "Реализовано", "ClamAV"),
            ("РСБ.1", "Регистрация событий безопасности", "+", "+", "+", "+", "Реализовано", "rsyslog"),
        ]

        for row_num, measure in enumerate(measures, start=2):
            for col_num, value in enumerate(measure, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 25

    # ============ DOCUMENT GENERATION ============

    def generate_document_pdf(
        self,
        template_content: str,
        form_data: dict[str, Any],
    ) -> bytes:
        """Generate a single document as PDF from template."""
        # Render template with form data
        template = self.jinja_env.from_string(template_content)
        html_content = template.render(**form_data)

        # Wrap in HTML document
        full_html = f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <style>
        body {{
            font-family: 'DejaVu Sans', 'Arial', sans-serif;
            font-size: 12pt;
            line-height: 1.5;
            margin: 2cm;
        }}
        h1 {{ text-align: center; font-size: 16pt; }}
        h2 {{ text-align: center; font-size: 14pt; }}
        h3 {{ font-size: 12pt; margin-top: 20pt; }}
        table {{ width: 100%; border-collapse: collapse; margin: 10pt 0; }}
        th, td {{ border: 1px solid #000; padding: 5pt; }}
        .signature {{ margin-top: 30pt; }}
    </style>
</head>
<body>
{html_content}
</body>
</html>
        """

        if not WEASYPRINT_AVAILABLE:
            return full_html.encode("utf-8")

        html = HTML(string=full_html)
        return html.write_pdf()


# API Endpoints for Reporting
from fastapi import APIRouter, HTTPException, Response
from pydantic import BaseModel

router = APIRouter()
report_generator = ComplianceReportGenerator()


class ReportRequest(BaseModel):
    company_id: str
    framework: str = "all"
    format: str = "pdf"  # pdf or excel


@router.post("/generate")
async def generate_report(request: ReportRequest):
    """Generate compliance report."""
    # Mock data - in production would fetch from database
    company_data = {
        "name": "ООО «Демо Компания»",
        "inn": "7707083893",
    }

    compliance_data = {
        "overall_score": 75,
        "total_checks": 50,
        "passed": 35,
        "warnings": 10,
        "failed": 5,
        "frameworks": [
            {"name": "152-ФЗ «О персональных данных»", "score": 78, "completed": 18, "total": 24},
            {"name": "187-ФЗ «О безопасности КИИ»", "score": 65, "completed": 10, "total": 16},
            {"name": "ГОСТ Р 57580", "score": 70, "completed": 22, "total": 32},
        ],
        "categories": [
            {
                "name": "Организационные меры",
                "items": [
                    {"check": "Назначен ответственный за ПДн", "status": "pass", "comment": "Приказ № 1"},
                    {"check": "Утверждена Политика ПДн", "status": "warning", "comment": "Ожидает подписи"},
                ],
            },
        ],
        "recommendations": [
            {"priority": "high", "action": "Утвердить Политику обработки ПДн", "deadline": "25.01.2024"},
            {"priority": "medium", "action": "Завершить обучение сотрудников", "deadline": "15.02.2024"},
        ],
        "documents": [
            {"title": "Политика обработки ПДн", "status": "draft", "status_label": "Черновик", "date": "15.01.2024"},
            {"title": "Приказ о назначении ответственного", "status": "approved", "status_label": "Утверждён", "date": "10.01.2024"},
        ],
    }

    if request.format == "excel":
        content = report_generator.generate_compliance_report_excel(
            company_data, compliance_data, request.framework
        )
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=compliance_report_{request.company_id}.xlsx"},
        )
    else:
        content = report_generator.generate_compliance_report_pdf(
            company_data, compliance_data, request.framework
        )
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=compliance_report_{request.company_id}.pdf"},
        )


@router.post("/document/{template_code}")
async def generate_document(template_code: str, form_data: dict[str, Any]):
    """Generate a single document from template."""
    # In production would fetch template from database
    template_content = """
    <h1>{{title}}</h1>
    <p>Организация: {{company_name}}</p>
    <p>ИНН: {{inn}}</p>
    """

    content = report_generator.generate_document_pdf(template_content, form_data)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={template_code}.pdf"},
    )
