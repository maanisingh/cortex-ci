"""
Multi-Format Export Service (Phase 4.3)
Supports export to CSV, Excel, JSON, and PDF formats.
"""

import csv
import io
import json
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any
from uuid import UUID

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import AuditLog, Constraint, Entity

logger = structlog.get_logger()


class ExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"
    PDF = "pdf"


class ExportService:
    """
    Service for exporting data in multiple formats.
    """

    def __init__(self, export_dir: str | None = None):
        self.export_dir = Path(export_dir or "/tmp/cortex-exports")
        self.export_dir.mkdir(parents=True, exist_ok=True)

    async def export_entities(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        format: ExportFormat,
        entity_type: str | None = None,
        include_risks: bool = False,
    ) -> dict[str, Any]:
        """
        Export entities in the specified format.

        Args:
            db: Database session
            tenant_id: Tenant ID
            format: Export format
            entity_type: Filter by entity type
            include_risks: Include risk scores in export

        Returns:
            Export result with file path or data
        """
        query = select(Entity).where(
            Entity.tenant_id == tenant_id,
            Entity.is_active == True,  # noqa: E712
        )
        if entity_type:
            query = query.where(Entity.type == entity_type)

        result = await db.execute(query)
        entities = result.scalars().all()

        data = []
        for entity in entities:
            row = {
                "id": str(entity.id),
                "name": entity.name,
                "type": entity.type.value if entity.type else None,
                "country": entity.country,
                "description": entity.description,
                "aliases": ", ".join(entity.aliases) if entity.aliases else "",
                "created_at": entity.created_at.isoformat() if entity.created_at else None,
                "updated_at": entity.updated_at.isoformat() if entity.updated_at else None,
            }

            if include_risks and entity.current_risk_score is not None:
                row["risk_score"] = entity.current_risk_score
                row["risk_level"] = entity.risk_level

            data.append(row)

        return await self._export_data(data, format, "entities")

    async def export_constraints(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        format: ExportFormat,
        constraint_type: str | None = None,
    ) -> dict[str, Any]:
        """
        Export constraints in the specified format.

        Args:
            db: Database session
            tenant_id: Tenant ID
            format: Export format
            constraint_type: Filter by constraint type

        Returns:
            Export result
        """
        query = select(Constraint).where(
            Constraint.tenant_id == tenant_id,
            Constraint.is_active == True,  # noqa: E712
        )
        if constraint_type:
            query = query.where(Constraint.type == constraint_type)

        result = await db.execute(query)
        constraints = result.scalars().all()

        data = []
        for constraint in constraints:
            data.append(
                {
                    "id": str(constraint.id),
                    "name": constraint.name,
                    "type": constraint.type.value if constraint.type else None,
                    "severity": constraint.severity.value if constraint.severity else None,
                    "description": constraint.description,
                    "source": constraint.source,
                    "effective_date": constraint.effective_date.isoformat()
                    if constraint.effective_date
                    else None,
                    "expiry_date": constraint.expiry_date.isoformat()
                    if constraint.expiry_date
                    else None,
                }
            )

        return await self._export_data(data, format, "constraints")

    async def export_audit_logs(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        format: ExportFormat,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> dict[str, Any]:
        """
        Export audit logs in the specified format.

        Args:
            db: Database session
            tenant_id: Tenant ID
            format: Export format
            start_date: Start of date range
            end_date: End of date range

        Returns:
            Export result
        """
        query = select(AuditLog).where(AuditLog.tenant_id == tenant_id)

        if start_date:
            query = query.where(AuditLog.created_at >= start_date)
        if end_date:
            query = query.where(AuditLog.created_at <= end_date)

        query = query.order_by(AuditLog.created_at.desc())

        result = await db.execute(query)
        logs = result.scalars().all()

        data = []
        for log in logs:
            data.append(
                {
                    "id": str(log.id),
                    "created_at": log.created_at.isoformat() if log.created_at else None,
                    "user_email": log.user_email,
                    "action": log.action.value if log.action else None,
                    "resource_type": log.resource_type,
                    "resource_id": str(log.resource_id) if log.resource_id else None,
                    "description": log.description,
                    "success": log.success,
                    "ip_address": str(log.ip_address) if log.ip_address else None,
                }
            )

        return await self._export_data(data, format, "audit_logs")

    async def export_risk_report(
        self,
        db: AsyncSession,
        tenant_id: UUID,
        format: ExportFormat,
    ) -> dict[str, Any]:
        """
        Export comprehensive risk report.

        Args:
            db: Database session
            tenant_id: Tenant ID
            format: Export format

        Returns:
            Export result
        """
        # Get entities with risk scores
        query = (
            select(Entity)
            .where(
                Entity.tenant_id == tenant_id,
                Entity.is_active == True,  # noqa: E712
                Entity.current_risk_score.isnot(None),
            )
            .order_by(Entity.current_risk_score.desc())
        )

        result = await db.execute(query)
        entities = result.scalars().all()

        data = []
        for entity in entities:
            data.append(
                {
                    "entity_id": str(entity.id),
                    "entity_name": entity.name,
                    "entity_type": entity.type.value if entity.type else None,
                    "country": entity.country,
                    "risk_score": entity.current_risk_score,
                    "risk_level": entity.risk_level,
                    "last_risk_update": entity.risk_score_updated_at.isoformat()
                    if entity.risk_score_updated_at
                    else None,
                }
            )

        return await self._export_data(data, format, "risk_report")

    async def _export_data(
        self,
        data: list[dict[str, Any]],
        format: ExportFormat,
        name: str,
    ) -> dict[str, Any]:
        """
        Export data in the specified format.

        Args:
            data: List of dictionaries to export
            format: Export format
            name: Base name for the export file

        Returns:
            Export result with content or file path
        """
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{name}_{timestamp}"

        if format == ExportFormat.CSV:
            return await self._to_csv(data, filename)
        elif format == ExportFormat.JSON:
            return await self._to_json(data, filename)
        elif format == ExportFormat.EXCEL:
            return await self._to_excel(data, filename)
        elif format == ExportFormat.PDF:
            return await self._to_pdf(data, filename, name)
        else:
            raise ValueError(f"Unsupported format: {format}")

    async def _to_csv(self, data: list[dict], filename: str) -> dict[str, Any]:
        """Convert data to CSV format."""
        if not data:
            return {
                "format": "csv",
                "count": 0,
                "content": "",
                "content_type": "text/csv",
            }

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return {
            "format": "csv",
            "count": len(data),
            "filename": f"{filename}.csv",
            "content": output.getvalue(),
            "content_type": "text/csv",
        }

    async def _to_json(self, data: list[dict], filename: str) -> dict[str, Any]:
        """Convert data to JSON format."""
        content = json.dumps(
            {
                "exported_at": datetime.utcnow().isoformat(),
                "count": len(data),
                "items": data,
            },
            indent=2,
            default=str,
        )

        return {
            "format": "json",
            "count": len(data),
            "filename": f"{filename}.json",
            "content": content,
            "content_type": "application/json",
        }

    async def _to_excel(self, data: list[dict], filename: str) -> dict[str, Any]:
        """Convert data to Excel format (using openpyxl if available)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Export"

            if data:
                # Header row
                headers = list(data[0].keys())
                header_font = Font(bold=True)
                header_fill = PatternFill(
                    start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                )

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Data rows
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return {
                "format": "xlsx",
                "count": len(data),
                "filename": f"{filename}.xlsx",
                "content": output.getvalue(),
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return await self._to_csv(data, filename)

    async def _to_pdf(self, data: list[dict], filename: str, title: str) -> dict[str, Any]:
        """Convert data to PDF format (using reportlab if available)."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))

            elements = []
            styles = getSampleStyleSheet()

            # Title
            elements.append(Paragraph(f"<b>{title.replace('_', ' ').title()}</b>", styles["Title"]))
            elements.append(Spacer(1, 12))
            elements.append(
                Paragraph(
                    f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                    styles["Normal"],
                )
            )
            elements.append(Paragraph(f"Total Records: {len(data)}", styles["Normal"]))
            elements.append(Spacer(1, 24))

            if data:
                # Create table
                headers = list(data[0].keys())
                table_data = [headers]
                for row in data[:100]:  # Limit to 100 rows for PDF
                    table_data.append([str(row.get(h, ""))[:50] for h in headers])

                table = Table(table_data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 8),
                            ("FONTSIZE", (0, 1), (-1, -1), 7),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.HexColor("#F0F0F0")],
                            ),
                        ]
                    )
                )
                elements.append(table)

                if len(data) > 100:
                    elements.append(Spacer(1, 12))
                    elements.append(
                        Paragraph(
                            f"<i>Note: Showing first 100 of {len(data)} records</i>",
                            styles["Normal"],
                        )
                    )

            doc.build(elements)
            output.seek(0)

            return {
                "format": "pdf",
                "count": len(data),
                "filename": f"{filename}.pdf",
                "content": output.getvalue(),
                "content_type": "application/pdf",
            }

        except ImportError:
            logger.warning("reportlab not installed, falling back to JSON")
            return await self._to_json(data, filename)


# Global instance
export_service = ExportService()
